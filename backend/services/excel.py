from __future__ import annotations

from pathlib import Path
from typing import Iterable


def rows_from_excel(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[index] or f"column_{index + 1}": value for index, value in enumerate(row)})
    return rows


def write_excel(path: Path, rows: Iterable[dict], sheet_name: str = "Export") -> Path:
    import xlsxwriter

    rows = list(rows)
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet(sheet_name[:31])
    headers = sorted({key for row in rows for key in row.keys()}) if rows else ["No Data"]
    header_format = workbook.add_format({"bold": True, "bg_color": "#DBEAFE", "border": 1})
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
        worksheet.set_column(col, col, 18)
    for row_index, row in enumerate(rows, start=1):
        for col, header in enumerate(headers):
            worksheet.write(row_index, col, row.get(header, ""))
    workbook.close()
    return path
